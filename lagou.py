# !/usr/bin/env python3.4
# encoding: utf-8

import json
import requests
import xlwt
import time
from openpyxl import Workbook


# 获取存储职位信息的json对象，遍历获得公司名、福利待遇、工作地点、学历要求、工作类型、发布时间、职位名称、薪资、工作年限
def get_json(url, datas):
    my_headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.119 Safari/537.36",
        "Referer": "https://www.lagou.com/jobs/list_Python?city=%E5%85%A8%E5%9B%BD&cl=false&fromSearch=true&labelWords=&suginput=",
        "Content-Type": "application/x-www-form-urlencoded;charset = UTF-8"
    }
    time.sleep(5)
    ses = requests.session()  # 获取session
    ses.headers.update(my_headers)  # 更新
    ses.get(
        "https://www.lagou.com/jobs/list_python?city=%E5%85%A8%E5%9B%BD&cl=false&fromSearch=true&labelWords=&suginput=")
    content = ses.post(url=url, data=datas)
    result = content.json()
    info = result['content']['positionResult']['result']
    info_list = []
    for job in info:
        information = []
        information.append(job['positionId'])  # 岗位对应ID
        information.append(job['city'])  # 岗位对应城市
        information.append(job['companyFullName'])  # 公司全名
        information.append(job['companyLabelList'])  # 福利待遇
        information.append(job['district'])  # 工作地点
        information.append(job['education'])  # 学历要求
        information.append(job['firstType'])  # 工作类型
        information.append(job['formatCreateTime'])  # 发布时间
        information.append(job['positionName'])  # 职位名称
        information.append(job['salary'])  # 薪资
        information.append(job['workYear'])  # 工作年限
        info_list.append(information)
        # 将列表对象进行json格式的编码转换,其中indent参数设置缩进值为2
        # print(json.dumps(info_list, ensure_ascii=False, indent=2))
    # print(info_list)
    return info_list


def main():
    page = int(input('请输入你要抓取的页码总数：'))
    kd = input('请输入你要抓取的职位关键字：')
    city = input('请输入你要抓取的城市：')

    info_result = []
    title = ['岗位id', '城市', '公司全名', '福利待遇', '工作地点', '学历要求', '工作类型', '发布时间', '职位名称', '薪资', '工作年限']
    info_result.append(title)
    for x in range(1, page + 1):
        url = 'https://www.lagou.com/jobs/positionAjax.json?needAddtionalResult=false'
        datas = {
            'first': 'false',
            'pn': x,
            'kd': kd,
            'city': city,
        }
        try:
            info = get_json(url, datas)
            info_result = info_result + info
            # print(info_result)
            print("第%s页正常采集" % x)
        except Exception as msg:
            print("第%s页出现问题" % x)
            print(msg)

        # # 创建workbook,即excel
        # workbook = xlwt.Workbook(encoding='utf-8')
        # # 创建表,第二参数用于确认同一个cell单元是否可以重设值
        # worksheet = workbook.add_sheet('lagouzp', cell_overwrite_ok=True)
        # for i, row in enumerate(info_result):
        #     # print(row)
        #     for j, col in enumerate(row):
        #         # print(col)
        #         worksheet.write(i, j, col)
        # workbook.save('lagouzp.xls')

        # openpyxl创建workbook, 即excel
        f = Workbook()
        worksheet = f.create_sheet('lagouzp', index=0)
        for i, row in enumerate(info_result):
            # print(row)
            for j, col in enumerate(row):
                # print(type(str((col))))
                colstr = str(col)
                worksheet.cell(row=i+1, column=j+1, value=colstr)
        f.save(city + kd + '--lagouzp.xlsx')


if __name__ == '__main__':
    main()
